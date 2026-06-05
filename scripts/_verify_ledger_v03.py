# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path

root = Path(__file__).resolve().parents[1]
v3 = root / "docs" / "story_database" / "02_前100篇原理去重台账_V0.3.xlsx"
wb = openpyxl.load_workbook(v3, data_only=True)
print("Sheets:", wb.sheetnames)
ws = wb["Vol1_LOCK"]
print("Vol1_LOCK rows:", ws.max_row)
ws2 = wb["前100篇台账"]
print("Total rows:", ws2.max_row)
print("Row2:", ws2.cell(2, 1).value, ws2.cell(2, 4).value)
print("Row7 (#1):", ws2.cell(7, 1).value, ws2.cell(7, 2).value, ws2.cell(7, 4).value)
print("Row56 (#50):", ws2.cell(56, 1).value, ws2.cell(56, 4).value)
print("Row57 (#51):", ws2.cell(57, 1).value, ws2.cell(57, 4).value)
print("Row106 (#100):", ws2.cell(106, 1).value, ws2.cell(106, 2).value, ws2.cell(106, 4).value)
