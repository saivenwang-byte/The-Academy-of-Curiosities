# -*- coding: utf-8 -*-
import glob, json
from pathlib import Path
import openpyxl

paths = glob.glob(str(Path(__file__).resolve().parents[1] / "【CC】files/files20260605-1_extracted/*.xlsx"))
wb = openpyxl.load_workbook(paths[0], data_only=True)
out = Path(__file__).resolve().parents[1] / "【CC】files/files20260605-1_extracted/_ledger_export.json"
data = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = []
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in row):
            rows.append(row)
    data[sn] = rows
out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
print("wrote", out)
