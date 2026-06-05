# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path

root = Path(__file__).resolve().parents[1]
v2 = next(root.glob("docs/story_database/02_*V0.2.xlsx"))
wb = openpyxl.load_workbook(v2, data_only=True)
ws = wb["Vol1_LOCK"]
print("Vol1_LOCK rows:", ws.max_row)
for r in range(1, ws.max_row + 1):
    print([ws.cell(r, c).value for c in range(1, 7)])

ws2 = wb["前50篇台账"]
print("\nRow2 canon:", ws2.cell(2, 1).value, ws2.cell(2, 2).value)
print("Row7 (#1 wet):", ws2.cell(7, 1).value, ws2.cell(7, 2).value, ws2.cell(7, 4).value)
