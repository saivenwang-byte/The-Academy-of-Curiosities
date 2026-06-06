# -*- coding: utf-8 -*-
"""Build 前50篇原理去重台账 V0.2 with Vol1 LOCK rows + canon_id/gate_pass."""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
CC = ROOT / "【CC】files" / "files20260605-1_extracted"


def resolve_source_xlsx(cc_dir: Path) -> Path:
    if not cc_dir.is_dir():
        raise FileNotFoundError(f"Source directory missing: {cc_dir}")
    matches = sorted(cc_dir.glob("*.xlsx"))
    if not matches:
        raise FileNotFoundError(f"No .xlsx found in {cc_dir}")
    if len(matches) > 1:
        print(f"WARN: multiple xlsx in {cc_dir.name}, using {matches[0].name}", file=sys.stderr)
    return matches[0]


SRC = resolve_source_xlsx(CC)
DST_V01 = ROOT / "docs" / "story_database" / "02_前50篇原理去重台账_V0.1.xlsx"
DST_V02 = ROOT / "docs" / "story_database" / "02_前50篇原理去重台账_V0.2.xlsx"
SHEET = "前50篇台账"

# Vol1 LOCK rows (insert before plan #1)
VOL1_ROWS = [
    [
        "A001",
        "ポスターの端",
        "翘边的海报",
        "I 组建期",
        "核心连载",
        "校园",
        "4月",
        "慧美·志郎·光",
        "光·慧美·志郎·珣（+瑆 笔记）",
        "气象",
        "材料",
        "流体",
        "侧廊出风 + 晨间湿度 + 胶带黏着",
        "实验+观察",
        "Vol1·样章",
        "G1–G7 Y",
    ],
    [
        "A002",
        "ずれた泥の跡",
        "错位的泥印",
        "I 组建期",
        "核心连载",
        "校园",
        "4月",
        "慧美·光",
        "光·慧美·志郎·珣（+瑆 笔记）",
        "行为",
        "动线",
        "地理",
        "鞋柜动线 + 目击链 + 泥印方向",
        "采访+实验",
        None,
        "G1–G7 Y",
    ],
    [
        "A003",
        "空いた欄",
        "空着的那一栏",
        "I 组建期",
        "核心连载",
        "校园",
        "4月",
        "慧美",
        "光·慧美·志郎·珣（+瑆 笔记）",
        "逻辑",
        "心理",
        "传播",
        "核实流程 + 采访伦理 + 无源收束",
        "采访+逻辑",
        None,
        "G1–G7 Y",
    ],
    [
        "A004",
        "チョークの円",
        "粉笔灰的圆圈",
        "I 组建期",
        "核心连载",
        "校园",
        "4月",
        "志郎·珣",
        "光·慧美·志郎·珣（+瑆 笔记）",
        "静电",
        "材料",
        "行为",
        "静电吸附 + 粉笔粉尘 + 清扫动线",
        "实验+结构",
        None,
        "G1–G7 Y",
    ],
    [
        "A005",
        "消しゴムの屑",
        "橡皮屑的方向",
        "I 组建期",
        "核心连载",
        "校园",
        "4月",
        "志郎·珣",
        "光·慧美·志郎·珣（+瑆 笔记）",
        "力学",
        "材料",
        "行为",
        "摩擦痕迹 + 振动滑出 + 行为链",
        "结构还原",
        "珣入社",
        "G1–G7 Y",
    ],
]

# Plan # -> canon_id remap (plan numbering in original sheet col A)
PLAN_REMAP = {
    1: ("C001", "G1=N · 素材库"),
    2: ("A006", "⬜"),
    3: ("A007", "⬜"),
    4: ("A008", "⬜"),
    5: ("A009", "⬜"),
    6: ("A010", "⬜"),
    7: ("A011", "⬜"),
    8: ("A012", "⬜"),
    9: ("A013", "⬜"),
    10: ("A014", "⬜"),
    11: ("A015", "⬜"),
    12: ("A016", "⬜"),
    13: ("A017", "⬜ · ★L01"),
    14: ("A018", "⬜"),
    15: ("A019", "⬜"),
    16: ("A020", "⬜"),
    17: ("A021", "⬜"),
    18: ("A022", "⬜"),
    19: ("A023", "⬜"),
    20: ("A024", "⬜"),
    21: ("B001", "⬜"),
    22: ("B002", "⬜"),
    23: ("B003", "⬜"),
    24: ("B004", "⬜"),
    25: ("B005", "⬜"),
    26: ("B006", "⬜"),
    27: ("B007", "⬜ · ★"),
    28: ("B008", "⬜"),
    29: ("B009", "⬜"),
    30: ("B010", "⬜ · ★"),
    31: ("A025", "⬜"),
    32: ("A026", "⬜"),
    33: ("A027", "⬜"),
    34: ("A028", "⬜"),
    35: ("C002", "⬜"),
    36: ("A029", "⬜"),
    37: ("A030", "⬜"),
    38: ("A031", "⬜"),
    39: ("A032", "⬜"),
    40: ("A033", "⬜"),
    41: ("A051", "⬜ · 理紗Vol11"),
    42: ("A052", "⬜"),
    43: ("A053", "⬜"),
    44: ("A054", "⬜"),
    45: ("A055", "⬜"),
    46: ("A056", "⬜"),
    47: ("A057", "⬜"),
    48: ("A058", "⬜"),
    49: ("B011", "⬜"),
    50: ("A059", "⬜ · ★收束1"),
}


def main() -> None:
    DST_V01.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC, DST_V01)

    wb = openpyxl.load_workbook(SRC)
    if SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet {SHEET!r} not in {SRC.name}; found: {wb.sheetnames}")
    ws = wb[SHEET]

    # Insert header columns at front
    ws.insert_cols(1, 2)
    ws.cell(1, 1, "正典ID")
    ws.cell(1, 2, "gate_pass")
    for c in (1, 2):
        ws.cell(1, c).font = Font(bold=True)

    # Shift: old col1 案号 is now col3
    old_max = ws.max_row

    # Insert 5 Vol1 rows after header
    ws.insert_rows(2, 5)
    for i, row in enumerate(VOL1_ROWS):
        r = 2 + i
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor="FFF2CC")

    # Update plan rows (now start at row 7 = old row 2 + 5 inserted)
    for r in range(7, old_max + 6):
        plan_num = ws.cell(r, 3).value  # 案号 column
        if isinstance(plan_num, (int, float)):
            plan_num = int(plan_num)
            cid, gate = PLAN_REMAP.get(plan_num, ("TBD", "⬜"))
            ws.cell(r, 1, cid)
            ws.cell(r, 2, gate)
            if plan_num == 1:
                ws.cell(r, 16, "C001·热学深潜")  # 长线标记 col shifted +2 -> was 15, now 17? 
                # After insert cols: 长线标记 was col 15, now 17
                ws.cell(r, 17, "C001·热学深潜")
                ws.cell(r, 17).fill = PatternFill("solid", fgColor="FCE4D6")

    # Fix row 7 (#1 wet chair) long line marker - find by 案号=1
    for r in range(7, ws.max_row + 1):
        if ws.cell(r, 3).value == 1:
            ws.cell(r, 17, "C001·热学深潜（非入社）")

    # Add Vol1_LOCK sheet
    if "Vol1_LOCK" in wb.sheetnames:
        del wb["Vol1_LOCK"]
    v1 = wb.create_sheet("Vol1_LOCK", 0)
    headers = [
        "正典ID", "gate_pass", "篇名（中）", "三原理三元组", "carrier", "status",
    ]
    v1.append(headers)
    statuses = ["样章+正文✅", "正文✅", "正文✅", "Card✅", "Card✅"]
    for i, row in enumerate(VOL1_ROWS):
        v1.append([
            row[0], row[15], row[2], row[12], row[5], statuses[i],
        ])

    # Update 说明 sheet
    if "说明" in wb.sheetnames:
        sh = wb["说明"]
        sh.cell(sh.max_row + 2, 1, "V0.2 · 2026-06-05")
        sh.cell(sh.max_row, 1, "· 增列 正典ID / gate_pass · 插入 Vol1 A001–A005 · #1→C001")

    wb.save(DST_V02)
    print("V0.1:", DST_V01)
    print("V0.2:", DST_V02)


if __name__ == "__main__":
    try:
        main()
    except (FileNotFoundError, KeyError, OSError) as e:
        print(f"FAIL: {e}", file=sys.stderr)
        sys.exit(1)
